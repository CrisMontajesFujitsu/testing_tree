import concurrent.futures
import subprocess
import tempfile

from openpyxl import load_workbook

from lib import ROW_LIMIT


# function that will execute tree command, uses ThreadPoolExecutor
# to allow concurrent execution of tree on Disc C:/ and Disc D:/
def run_tree(command):
    def worker():
        with tempfile.TemporaryDirectory() as cache_dir:
            result = subprocess.run(command, stdout=subprocess.PIPE, text=True, shell=True, cwd=cache_dir)
            if result.returncode == 0:
                return result.stdout

    with concurrent.futures.ThreadPoolExecutor() as exec:
        tree = exec.submit(worker)

    return None if "Invalid drive specification" in tree.result() else tree.result()

# function to write the result of tree command to passed workbook(excel)
def write_to_xl(items, sheet_name_prefix, workbook):
    cp_file = items.copy()
    sheets = []
    sheet_page = 0
    last_sheet_index = len(workbook.worksheets) - 1
    active_sheet = None

    for i, line in enumerate(cp_file):
        if i % ROW_LIMIT == 0:
            sheet_page += 1
            sheet_name = f"{sheet_name_prefix}_{sheet_page}"
            sheets.append(sheet_name)
            sheets[sheet_page - 1] = workbook.create_sheet(title=sheet_name, index=last_sheet_index)
            active_sheet = sheets[sheet_page - 1]

        active_sheet.append([str(line)])

# write the result of both tree command on Disk C:/ and Disk D:/ to final LCF file
def tree_to_LCF(tree_c, tree_d, name):
    print(type(tree_d.result()))

    # load LCF template
    workbook = load_workbook(filename="test.xlsx", read_only=False)

    # write result of tree command on Drive C to LCF
    # write_to_xl(tree_c.result().split('\n'), "Drive C", workbook)

    # check first if user has Drive D before writing to LCF
    if tree_d.result() is not None:
        write_to_xl(tree_d.result().split('\n'), "Drive D", workbook)

    # save
    workbook.save(f'test_res_{name}.xlsx')


##########
def main():
    with concurrent.futures.ThreadPoolExecutor() as exec:
        # tree_c = exec.submit(run_tree, "tree /f C:\\")
        tree_d = exec.submit(run_tree, "tree /f D:\\")
    
    tree_to_LCF("tree_c", tree_d, "Testing")


main()
